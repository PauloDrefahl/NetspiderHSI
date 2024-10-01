import pandas as pd
import os
from datetime import datetime
from shutil import copy2
from PyPDF2 import PdfMerger
import openpyxl
from openpyxl.utils import get_column_letter


class FolderAppender:
    def __init__(self, results_directory):
        self.new_folder_name = None
        self.new_file_path = None
        self.new_file_name = None
        self.new_screenshot_folder_path = None
        self.new_folder_path = None
        self.results_directory = results_directory
        self.selected_folders = None
        self.pdf_merger = PdfMerger()
        self.dataframes = []

    def setSelectedFolders(self, selected_folders):
        self.selected_folders = selected_folders

    def create_new_folder(self):
        self.new_folder_name = '-'.join(folder.split('-')[0] for folder in self.selected_folders)
        self.new_folder_name += '-' + '-'.join(folder.split('-')[1] for folder in self.selected_folders)
        self.new_folder_name += '-' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

        self.new_folder_path = os.path.join(self.results_directory, self.new_folder_name)
        os.makedirs(self.new_folder_path, exist_ok=True)

        self.new_screenshot_folder_path = os.path.join(self.new_folder_path, 'screenshots')
        os.makedirs(self.new_screenshot_folder_path, exist_ok=True)

        self.new_file_name = 'CLEAN-' + self.new_folder_name
        self.new_file_path = os.path.join(self.new_folder_path, self.new_file_name + '.xlsx')

    def append_files(self):
        folder_count = 0
        for folder in self.selected_folders:
            folder_path = os.path.join(self.results_directory, folder)
            spreadsheet_name = 'CLEAN-' + folder + '.xlsx'
            file_path = os.path.join(folder_path, spreadsheet_name)
            if os.path.exists(file_path):
                df = pd.read_excel(file_path)
                self.dataframes.append(df)

            screenshots_path = os.path.join(folder_path, 'screenshots')
            if os.path.exists(screenshots_path):
                for i, file in enumerate(os.listdir(screenshots_path)):
                    if file.endswith('.png'):
                        new_screenshot_file_name = f"{folder_count}_{i}.png"
                        source_file_path = os.path.join(screenshots_path, file)
                        destination_file_path = os.path.join(self.new_screenshot_folder_path, new_screenshot_file_name)
                        copy2(source_file_path, destination_file_path)

            folder_count += 1

            pdf_path = os.path.join(folder_path, 'screenshots')
            if os.path.exists(pdf_path):
                for file in os.listdir(pdf_path):
                    if file.endswith('.pdf'):
                        self.pdf_merger.append(os.path.join(pdf_path, file))

    def save_data(self):
        combined_df = pd.concat(self.dataframes, ignore_index=True)
        combined_df.fillna('N/A', inplace=True)
        combined_df['Post-identifier'] = range(len(combined_df))
        combined_df.to_excel(self.new_file_path, index=False)

        workbook = openpyxl.load_workbook(self.new_file_path)
        worksheet = workbook.active

        for col in worksheet.columns:
            max_length = 0
            column = col[0].column
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except Exception as e:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

        workbook.save(self.new_file_path)
        workbook.close()

        merged_pdf_path = os.path.join(self.new_screenshot_folder_path, self.new_folder_name + '.pdf')
        self.pdf_merger.write(merged_pdf_path)
        self.pdf_merger.close()
