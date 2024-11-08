import os
from datetime import datetime
import pandas as pd
from seleniumbase import Driver
from seleniumbase.common.exceptions import NoSuchElementException
from selenium.webdriver.common.by import By
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill


class SkipthegamesScraper(ScraperPrototype):
    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None
        self.cities = {
            "bonita springs": 'https://skipthegames.com/posts/bonita-springs-fl',
            "bradenton": 'https://skipthegames.com/posts/bradenton',
            "cape coral": 'https://skipthegames.com/posts/cape-coral-fl',
            "fort myers": 'https://skipthegames.com/posts/fort-myers',
            "ocala": 'https://skipthegames.com/posts/ocala',
            "okaloosa": 'https://skipthegames.com/posts/okaloosa',
            "orlando": 'https://skipthegames.com/posts/orlando',
            "palm bay": 'https://skipthegames.com/posts/palmbay',
            "gainesville": 'https://skipthegames.com/posts/gainesville',
            "jacksonville": 'https://skipthegames.com/posts/jacksonville',
            "keys": 'https://skipthegames.com/posts/keys',
            "miami": 'https://skipthegames.com/posts/miami',
            "naples": 'https://skipthegames.com/posts/naples-fl',
            "st. augustine": 'https://skipthegames.com/posts/st-augustine',
            "tallahassee": 'https://skipthegames.com/posts/tallahassee',
            "tampa": 'https://skipthegames.com/posts/tampa',
            "sarasota": 'https://skipthegames.com/posts/sarasota',
            "space coast": 'https://skipthegames.com/posts/space-coast',
            "venice": 'https://skipthegames.com/posts/venice-fl',
            "west palm beach": 'https://skipthegames.com/posts/west-palm-beach'
        }
        self.city = ''
        self.url = ''

        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.keywords = None
        self.flagged_keywords = None
        self.completed = False

        self.join_keywords = False
        self.search_mode = False
        self.number_of_keywords_in_post = 0
        self.keywords_found_in_post = []

        # lists to store data and then send to excel file
        self.link = []
        self.about_info = []
        self.description = []
        self.services = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.only_posts_with_payment_methods = False

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''
    def initialize(self) -> None:
        # set keywords value
        #self.keywords = keywords

        # set up directories to save screenshots and excel file.
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        # Selenium Web Driver setup
        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()

        # Find links of posts
        links = self.get_links()

        # create directories for screenshot and excel
        self.scraper_directory = f'{self.path}/skipthegames-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/skipthegames-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)

        self.get_data(links)
        self.close_webpage()
        self.reset_variables()

    def stop_scraper(self) -> None:
        if self.search_mode:
            self.driver.close()
            self.driver.quit()
        else:
            self.driver.close()

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.close()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> set:
        posts = self.driver.find_elements(
            By.CSS_SELECTOR, 'html.no-js body div table.two-col-wrap tbody tr '
                             'td#gallery_view.listings-with-sidebar.list-search-results.gallery div.full-width '
                             'div.small-16.columns div.clsfds-display-mode.gallery div.day-gallery [href]')
        links = [post.get_attribute('href') for post in posts]

        # remove sponsored links
        links = [link for link in links if link.startswith('https://skipthegames.com/posts/')]
        return set(links)

    def get_formatted_url(self) -> None:
        self.url = self.cities.get(self.city)

    def get_data(self, links) -> None:
        counter = 0

        for link in links:
            self.driver.get(link)
            assert "Page not found" not in self.driver.page_source
            try:
                about_info = self.driver.wait_for_element(
                    By.CSS_SELECTOR, '#post-body tbody').text
            except NoSuchElementException:
                about_info = 'N/A'

            try:
                services = self.driver.wait_for_element(
                    By.XPATH, '//*[@id="post-services"]').text
            except NoSuchElementException:
                services = 'N/A'

            try:
                # NOTE: It's possible for two elements to have the same ID.
                description = self.driver.wait_for_element(
                    By.CSS_SELECTOR, '#post-body #post-body').text
            except NoSuchElementException:
                description = 'N/A'

            # reassign variables for each post
            self.number_of_keywords_in_post = 0
            self.keywords_found_in_post = []

            if self.join_keywords and self.only_posts_with_payment_methods:
                if self.check_keywords(about_info) or self.check_keywords(services) or self.check_keywords(description):
                    self.check_keywords_found(about_info, services, description, link)
                    counter = self.join_with_payment_methods(about_info, counter, description, link, services)

            elif self.join_keywords or self.only_posts_with_payment_methods:
                if self.join_keywords:
                    if self.check_keywords(about_info) or self.check_keywords(services) or self.check_keywords(
                            description):
                        self.check_keywords_found(about_info, services, description, link)
                        counter = self.join_inclusive(about_info, counter, description, link, services)

                elif self.only_posts_with_payment_methods:
                    if len(self.keywords) > 0:
                        if self.check_keywords(about_info) or self.check_keywords(services) or self.check_keywords(
                                description):
                            self.check_keywords_found(about_info, services, description, link)
                    counter = self.payment_methods_only(about_info, counter, description, link, services)
            else:
                if len(self.keywords) > 0:
                    if self.check_keywords(about_info) or self.check_keywords(services) or self.check_keywords(
                            description):
                        self.check_keywords_found(about_info, services, description, link)
                        self.append_data(about_info, counter, description, link, services)
                        screenshot_name = str(counter) + ".png"
                        self.capture_screenshot(screenshot_name)
                        counter += 1
                else:
                    self.append_data(about_info, counter, description, link, services)
                    screenshot_name = str(counter) + ".png"
                    self.capture_screenshot(screenshot_name)
                    counter += 1

            self.RAW_format_data_to_excel()
            self.CLEAN_format_data_to_excel()

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, about_info, counter, description, link, services):
        self.post_identifier.append(counter)
        self.link.append(link)
        self.about_info.append(about_info)
        self.services.append(services)
        self.description.append(description)
        self.check_and_append_payment_methods(description)
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(self.number_of_keywords_in_post or 'N/A')
        self.check_for_social_media(description)

    def join_with_payment_methods(self, about_info, counter, description, link, services) -> int:
        if self.check_for_payment_methods(description) and len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(about_info, counter, description, link, services)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def join_inclusive(self, about_info, counter, description, link, services) -> int:
        if len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(about_info, counter, description, link, services)

            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def payment_methods_only(self, about_info, counter, description, link, services) -> int:

        if self.check_for_payment_methods(description):
            self.append_data(about_info, counter, description, link, services)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, about_info, services, description, link) -> None:
        self.check_and_append_keywords(about_info)
        self.check_and_append_keywords(services)
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(link)

    def check_for_payment_methods(self, description) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def check_and_append_payment_methods(self, description) -> None:
        payments = ''
        for payment in self.known_payment_methods:
            if payment in description.lower():
                payments += payment + '\n'

        if payments != '':
            self.payment_methods_found.append(payments)
        else:
            self.payment_methods_found.append('N/A')

    def check_for_social_media(self, description) -> None:
        social_media = ''
        for social in self.known_social_media:
            if social in description.lower():
                social_media += social + '\n'

        if social_media != '':
            self.social_media_found.append(social_media)
        else:
            self.social_media_found.append('N/A')

    def check_keywords(self, data) -> bool:
        for key in self.keywords:
            if key in data:
                return True
        return False

    def check_and_append_keywords(self, data) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.append(key)
                self.number_of_keywords_in_post += 1

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = {
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            'Inputted City / Region': self.city,
            # -------
            'About-info': self.about_info,
            'Description': self.description,
            'Services': self.services,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        }

        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-skipthegames-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']

            for i in range(2, worksheet.max_row):
                keywords = worksheet["G" + str(
                    i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["G" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        personal_info = [
            f"{ab_info } "
            for ab_info in zip(
                self.about_info
            )
        ]

        overall_desc = [
            f"{description} ||| {services} "
            for description, services in zip(
                self.description, self.services
            )
        ]

        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # ------- use the setter attributes to add into the spreadsheet for location
            'Inputted City / Region': self.city,
            'Specified Location': 'N/A',
            'Timeline': 'N/A',
            'Contacts': 'N/A',
            'Personal Info': personal_info,
            'Overall Description': overall_desc,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-skipthegames-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def capture_screenshot(self, screenshot_name) -> None:
        self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
        self.create_pdf()

    def create_pdf(self) -> None:
        screenshot_files = [os.path.join(self.screenshot_directory, filename) for filename in
                            os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.link = []
        self.about_info = []
        self.description = []
        self.services = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.social_media_found = []
