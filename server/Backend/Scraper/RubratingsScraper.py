import os
import time
import urllib.parse
from datetime import datetime
import pandas as pd
from seleniumbase import Driver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from typing_extensions import override
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill


class RubratingsScraper(ScraperPrototype):

    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None
        self.cities = {
            "fort myers": "http://rubsratings.com/list/list/39",
            "gainesville": "http://rubsratings.com/list/list/40",
            "jacksonville": "http://rubsratings.com/list/list/41",
            "miami": "http://rubsratings.com/list/list/42",
            "ft lauderdale": "http://rubsratings.com/list/list/43",
            "orlando": "http://rubsratings.com/list/list/44",
            "panama city": "http://rubsratings.com/list/list/45",
            "pensacola": "http://rubsratings.com/list/list/46",
            "tallahassee": "http://rubsratings.com/list/list/47",
            "tampa": "http://rubsratings.com/list/list/48",
        }
        self.city = ''
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.pdf = None
        self.keywords: set[str] = set()
        self.flagged_keywords: set[str] = set()
        self.search_mode = False
        self.completed = False

        self.join_keywords = False
        self.keywords_found_in_post: set[str] = set()

        self.only_posts_with_payment_methods = False

        # lists to store data and then send to csv file
        self.last_activity = []
        self.phone_number = []
        self.link = []
        self.location = []
        self.provider_id = []
        self.post_title = []
        self.description = []
        self.post_identifier = []
        self.payment_methods_found = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords: set[str]) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords: set[str]) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''
    def initialize(self) -> None:
        # set up directories to save screenshots and Excel file.
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            # window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()

        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/rubratings-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/rubratings-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)
        self.get_data(links)
        self.close_webpage()
        self.reset_variables()
        self.completed = True

    def stop_scraper(self) -> None:
        self.completed = True

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        if self.search_mode:
            self.driver.get(self.url)
        else:
            self.driver.execute_script(f'window.open("{self.url}", "_blank");')
            original_window = self.driver.current_window_handle
            time.sleep(5)
            if len(self.driver.window_handles) >= 2:
                # Switch to the new tab
                self.driver.switch_to.window(self.driver.window_handles[-1])

                # Close the original tab
                self.driver.switch_to.window(original_window)
                self.driver.close()

                # Switch back to the new tab
                self.driver.switch_to.window(self.driver.window_handles[-1])

        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.quit()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> list[str]:
        links = self.driver.find_elements(By.CSS_SELECTOR, ".container .list-item a")
        urls = [link.get_attribute("href") for link in links]
        return urls

    def get_formatted_url(self) -> None:
        self.url = self.cities.get(self.city)

    def get_data(self, links: list[str]) -> None:
        counter = 1

        for link in links:
            print(f"Processing link {counter}/{len(links)}: {link}")
            if not self.completed:
                self.driver.implicitly_wait(10)
                self.driver.get(link)
                time.sleep(1)
                self.driver.get(link)
                assert "Page not found" not in self.driver.page_source
                try:
                    label = "Latest Activity: "
                    # NOTE: XPath 1.0 does not allow quotes to be escaped.
                    assert "'" not in label
                    last_activity = self.driver.find_element(
                        By.XPATH, f"//*[contains(text(), '{label}')]"
                    ).text.replace(label, "")
                except NoSuchElementException:
                    last_activity = 'N/A'
                try:
                    phone_number_element = self.driver.find_element(By.XPATH, '/html/body/div[2]/div[3]/div/div['
                                                                            '1]/div/div/div/div[2]/div[3]/div[1]/ul/li['
                                                                            '1]/a')
                    phone_number = phone_number_element.get_attribute('data-replace')
                except NoSuchElementException:
                    phone_number = 'N/A'

                try:
                    label = "Location: "
                    # NOTE: XPath 1.0 does not allow quotes to be escaped.
                    assert "'" not in label
                    location = self.driver.find_element(
                        By.XPATH, f"//li[contains(., '{label}')]"
                    ).text.replace(label, "")
                except NoSuchElementException:
                    location = 'N/A'

                path = urllib.parse.urlsplit(link).path
                # The last segment is always the provider ID.
                provider_id = "#" + path.strip("/").rsplit("/", 1)[-1].strip()

                try:
                    post_title = self.driver.find_element(
                        By.CSS_SELECTOR, "#provider .info-outer h3"
                    ).text
                except NoSuchElementException:
                    post_title = 'N/A'
                try:
                    # Find the first element with the specified XPath
                    first_element = self.driver.find_element(By.XPATH,
                                                            '/html/body/div[2]/div[3]/div/div[1]/div/div/div/div[2]/p[2] |'
                                                            '/html/body/div[2]/div[3]/div/div[1]/div/div/div/div[2]/h2[1]')

                    # Determine if the first element is an <h2> element
                    if first_element.tag_name == 'h2':
                        # If it's an <h2> element, find following sibling <h2> elements
                        following_elements = first_element.find_elements(By.XPATH, './following-sibling::h2')
                    else:
                        # If it's a <p> element, find following sibling <p> elements
                        following_elements = first_element.find_elements(By.XPATH, './following-sibling::p')

                    # Extract text from the first element
                    description_texts = [first_element.text]

                    # Extract text from all found elements
                    description_texts.extend(element.text for element in following_elements)

                    # Concatenate the descriptions into a single string
                    description = ' '.join(description_texts)

                except NoSuchElementException:
                    description = 'N/A'

                # reassign variables for each post
                self.keywords_found_in_post.clear()

                # Search the post's contents for keywords.
                self.check_keywords_found(last_activity, phone_number, location, provider_id, post_title, description, link)

                if self._should_discard_post(description):
                    continue

                # Save the data we collected about the post.
                self.append_data(counter, link, last_activity, phone_number, location, provider_id, post_title, description)
                screenshot_name = str(counter) + ".png"
                self.capture_screenshot(screenshot_name)
                counter += 1

                self.RAW_format_data_to_excel()
                self.CLEAN_format_data_to_excel()
            # Breaks the links loop for fast closing time once user presses stop scraper
            else:
                break

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, counter, link, last_activity, phone_number, location, provider_id, post_title, description):
        self.post_identifier.append(counter)
        self.link.append(link)
        self.last_activity.append(last_activity)
        self.phone_number.append(phone_number)
        self.location.append(location)
        self.provider_id.append(provider_id)
        self.post_title.append(post_title)
        self.description.append(description)
        payment_methods = self.get_payment_methods(description)
        self.payment_methods_found.append("\n".join(payment_methods) or "N/A")
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(len(self.keywords_found_in_post) or "N/A")
        social_media = self.get_social_media(description)
        self.social_media_found.append("\n".join(social_media) or "N/A")
        # Store information about the post in the database.
        try:
            with self.open_database() as connection, connection.cursor() as cursor:
                # PostgreSQL expects `last_activity` to contain the year, but
                # RubRatings only shows the month and day, e.g., 'Mon, 9 Dec'.
                last_activity += " " + str(datetime.now(tz=None).year)
                # In the database, the provider ID is stored as an `integer`, so
                # we must remove the '#' and cast `provider_id` to an `int`.
                provider_id = provider_id.removeprefix("#")
                # TODO(Daniel): Remove the check for "N/A" once the provider ID
                # locator is fixed; every page should have a provider ID.
                provider_id = None if provider_id == "N/A" else int(provider_id)
                cursor.execute(
                    """
                    insert into raw_rub_ratings_posts
                    values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict do nothing;
                    """,
                    (
                        link,
                        self.city,
                        location,
                        last_activity,
                        phone_number,
                        provider_id,
                        post_title or None,
                        description or None,
                        payment_methods,
                        social_media,
                        list(self.keywords_found_in_post),
                    ),
                )
        except Exception as e:
            print(f"Database write failed: {e}")

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, last_activity, phone_number, location, provider_id, post_title, description, link) -> None:
        self.check_and_append_keywords(last_activity)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(location)
        self.check_and_append_keywords(provider_id)
        self.check_and_append_keywords(post_title)
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(link)

    @override
    def check_for_payment_methods(self, description: str) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def get_payment_methods(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        payment_methods: list[str] = []
        for payment_method in self.known_payment_methods:
            if payment_method in description:
                payment_methods.append(payment_method)
        return payment_methods

    def get_social_media(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        social_media: list[str] = []
        for social in self.known_social_media:
            if social in description:
                social_media.append(social)
        return social_media

    @override
    def check_and_append_keywords(self, data: str) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.add(key)

    def _should_discard_post(self, description: str) -> bool:
        if self.join_keywords:
            # Discard posts that don't contain ALL keywords.
            if len(self.keywords_found_in_post) < len(self.keywords):
                return True
        elif not self.only_posts_with_payment_methods and len(self.keywords) > 0:
            # Discard posts that don't contain ANY keywords, unless:
            # 1. We're specifically looking for posts with payment methods, in
            #    which case we keep *all* posts with payment methods.
            # 2. No keywords were originally provided.
            if len(self.keywords_found_in_post) == 0:
                return True

        if self.only_posts_with_payment_methods:
            if not self.check_for_payment_methods(description):
                return True

        return False

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': self.location,
            'Last-Activity': self.last_activity,
            'Phone-Number': self.phone_number,
            'Provider-ID': self.provider_id,
            'Post-Title': self.post_title,
            'Description': self.description,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-rubratings-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["L" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["L" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        # Concatenate attributes to fit into the CLEAN spreadsheet format, which
        # is consistent across all scrapers.
        overall_description = [
            f"{post_title} ||| {description}"
            for (post_title, description) in zip(
                self.post_title, self.description, strict=True
            )
        ]

        # These timestamps do *not* indicate when the post was created, unlike
        # the timestamps from YesBackpage and Escort Alligator.
        timeline = [
            f"Last Activity: {last_activity}"
            for last_activity in self.last_activity
        ]

        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # ------
            'Inputted City / Region': self.city,
            'Specified Location': self.location,
            'Timeline': timeline,
            'Contacts': self.phone_number,
            'Personal Info': self.provider_id,
            'Overall Description': overall_description,
            # ------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-rubratings-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def capture_screenshot(self, screenshot_name) -> None:
        try:
            # Wait for the picture under the 'pic' class to be visible
            WebDriverWait(self.driver, 10).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, '.pic img'))
            )

            # Take screenshot after the picture is visible
            self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
            self.create_pdf()
        except Exception as e:
            print(f"Error capturing screenshot: {e}")

    def create_pdf(self) -> None:
        screenshot_files = [os.path.join(self.screenshot_directory, filename) for filename in
                            os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.last_activity = []
        self.phone_number = []
        self.location = []
        self.provider_id = []
        self.post_title = []
        self.description = []
        self.link = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.keywords_found = []
        self.social_media_found = []
