import os
import time
from datetime import datetime
import pandas as pd
from seleniumbase import Driver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from typing_extensions import override
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill


class MegapersonalsScraper(ScraperPrototype):

    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None
        self.cities = {
            "daytona": 'https://megapersonals.eu/public/post_list/109/1/1',
            "fort lauderdale": 'https://megapersonals.eu/public/post_list/113/1/1',
            "fort myers": 'https://megapersonals.eu/public/post_list/234/1/1',
            "gainesville": 'https://megapersonals.eu/public/post_list/235/1/1',
            "jacksonville": 'https://megapersonals.eu/public/post_list/236/1/1',
            "keys": 'https://megapersonals.eu/public/post_list/114/1/1',
            "miami": 'https://megapersonals.eu/public/post_list/25/1/1',
            "ocala": 'https://megapersonals.eu/public/post_list/238/1/1',
            "okaloosa": 'https://megapersonals.eu/public/post_list/239/1/1',
            "orlando": 'https://megapersonals.eu/public/post_list/18/1/1',
            "palm bay": 'https://megapersonals.eu/public/post_list/110/1/1',
            "panama city": 'https://megapersonals.eu/public/post_list/240/1/1',
            "pensacola": 'https://megapersonals.eu/public/post_list/241/1/1',
            "sarasota": 'https://megapersonals.eu/public/post_list/242/1/1',
            "space coast": 'https://megapersonals.eu/public/post_list/111/1/1',
            "st. augustine": 'https://megapersonals.eu/public/post_list/243/1/1',
            "tallahassee": 'https://megapersonals.eu/public/post_list/244/1/1',
            "tampa": 'https://megapersonals.eu/public/post_list/50/1/1',
            "treasure coast": 'https://megapersonals.eu/public/post_list/112/1/1',
            "west palm beach": 'https://megapersonals.eu/public/post_list/115/1/1'
        }
        self.city = ''
        self.url = "https://megapersonals.eu/"
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        # set date variables and path
        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.keywords: set[str] = set()
        self.flagged_keywords: set[str] = set()

        self.only_posts_with_payment_methods = False
        self.completed = False

        self.join_keywords = False
        self.search_mode = False
        self.keywords_found_in_post: set[str] = set()

        # lists to store data and then send to excel file
        self.description = []
        self.name = []
        self.phoneNumber = []
        self.contentCity = []
        self.location = []
        self.link = []
        self.post_identifier = []
        self.payment_methods_found = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords: set[str]) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords: set[str]) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''
    def initialize(self) -> None:
        # format date
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            # window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()

        # Format website URL based on state and city
        self.get_formatted_url()
        self.driver.get(self.url)

        # Find links of posts
        links = self.get_links()

        # create directories for screenshot and excel
        self.scraper_directory = f'{self.path}/megapersonals-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/megapersonals-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)

        self.get_data(links)
        self.close_webpage()
        self.reset_variables()
        self.completed = True

    def stop_scraper(self) -> None:
        self.completed = True

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source
        self.driver.find_element(By.XPATH, '//*[@id="checkbox-agree"]').click()
        # There is a quick fade-in during which we cannot click the button.
        # `driver.click()` automatically waits for the button to be interactive.
        self.driver.click(By.ID, 'ageagree')
        self.driver.find_element(By.XPATH, '//*[@id="choseCityContainer"]/div[3]/label').click()
        self.driver.find_element(By.XPATH, '//*[@id="choseCityContainer"]/div[3]/article/div[10]/label').click()
        self.driver.find_element(By.XPATH,
                                 '//*[@id="choseCityContainer"]/div[3]/article/div[10]/article/p[3]/a').click()
        self.driver.find_element(By.XPATH, '//*[@id="megapCategoriesOrangeButton"]').click()

    def close_webpage(self) -> None:
        self.driver.quit()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> set:
        post_list = self.driver.find_elements(By.CLASS_NAME, 'listadd')

        # traverse through list of people to grab page links
        links = []
        for person in post_list:
            links.append(person.find_element(By.TAG_NAME, "a").get_attribute("href"))
        return set(links)

    def get_formatted_url(self):
        self.url = self.cities.get(self.city)

    def get_data(self, links) -> None:
        counter = 1

        for link in links:
            print(f"Processing link {counter}/{len(links)}: {link}")
            if not self.completed:
                self.driver.get(link)
                time.sleep(1)
                self.driver.get(link)
                assert "Page not found" not in self.driver.page_source

                try:
                    description = self.driver.find_element(
                        By.XPATH, '/html/body/div/div[6]/span').text
                except NoSuchElementException:
                    description = 'N/A'

                try:
                    phone_number = self.driver.find_element(
                        By.XPATH, '/html/body/div/div[6]/div[1]/span').text
                except NoSuchElementException:
                    phone_number = 'N/A'

                try:
                    name = self.driver.find_element(
                        By.XPATH, '/html/body/div/div[6]/p[1]/span[2]').text[5:]
                except NoSuchElementException:
                    name = 'N/A'

                try:
                    city = self.driver.find_element(
                        By.XPATH, '/html/body/div/div[6]/p[1]/span[1]').text[5:]
                except NoSuchElementException:
                    city = 'N/A'

                try:
                    location = self.driver.find_element(
                        By.XPATH, '/html/body/div/div[6]/p[2]').text[9:]
                except NoSuchElementException:
                    location = 'N/A'

                # reassign variables for each post
                self.keywords_found_in_post.clear()

                # Search the post's contents for keywords.
                self.check_keywords_found(city, description, location, name, phone_number, link)

                if self._should_discard_post(description):
                    continue

                # Save the data we collected about the post.
                self.append_data(city, counter, description, link, location, name, phone_number)
                screenshot_name = str(counter) + ".png"
                self.capture_screenshot(screenshot_name)
                counter += 1

                self.RAW_format_data_to_excel()
                self.CLEAN_format_data_to_excel()
            # Breaks the links loop for fast closing time once user presses stop scraper
            else:
                break

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, city, counter, description, link, location, name, phone_number):
        self.post_identifier.append(counter)
        self.name.append(name)
        self.phoneNumber.append(phone_number)
        self.contentCity.append(city)
        self.location.append(location)
        self.description.append(description)
        payment_methods = self.get_payment_methods(description)
        self.payment_methods_found.append("\n".join(payment_methods) or "N/A")
        self.link.append(link)
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(len(self.keywords_found_in_post) or "N/A")
        social_media = self.get_social_media(description)
        self.social_media_found.append("\n".join(social_media) or "N/A")
        # Store information about the post in the database.
        try:
            with self.open_database() as connection, connection.cursor() as cursor:
                cursor.execute(
                    """
                    insert into raw_mega_personals_posts
                    values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict do nothing;
                    """,
                    (
                        link,
                        self.city,
                        city,
                        location,
                        phone_number,
                        name,
                        description,
                        payment_methods,
                        social_media,
                        list(self.keywords_found_in_post),
                    ),
                )
        except Exception as e:
            print(f"Database write failed: {e}") 
    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, city, description, location, name, phone_number, link) -> None:
        self.check_and_append_keywords(city)
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(location)
        self.check_and_append_keywords(name)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(link)

    @override
    def check_for_payment_methods(self, description: str) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def get_payment_methods(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        payment_methods: list[str] = []
        for payment_method in self.known_payment_methods:
            if payment_method in description:
                payment_methods.append(payment_method)
        return payment_methods

    def get_social_media(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        social_media: list[str] = []
        for social in self.known_social_media:
            if social in description:
                social_media.append(social)
        return social_media

    @override
    def check_and_append_keywords(self, data: str) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.add(key)

    def _should_discard_post(self, description: str) -> bool:
        if self.join_keywords:
            # Discard posts that don't contain ALL keywords.
            if len(self.keywords_found_in_post) < len(self.keywords):
                return True
        elif not self.only_posts_with_payment_methods and len(self.keywords) > 0:
            # Discard posts that don't contain ANY keywords, unless:
            # 1. We're specifically looking for posts with payment methods, in
            #    which case we keep *all* posts with payment methods.
            # 2. No keywords were originally provided.
            if len(self.keywords_found_in_post) == 0:
                return True

        if self.only_posts_with_payment_methods:
            if not self.check_for_payment_methods(description):
                return True

        return False

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = {
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # ------
            'Inputted City / Region': self.city,
            'Specified City / Region': self.contentCity,
            'Specified Location': self.location,
            'Phone-number': self.phoneNumber,
            'Name': self.name,
            'Description': self.description,
            # ------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        }

        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-megapersonals-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']

            for i in range(2, worksheet.max_row):
                keywords = worksheet["I" + str(
                    i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["I" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        # Concatenate attributes to fit into the CLEAN spreadsheet format, which
        # is consistent across all scrapers.
        specified_location = [
            f"{contentCity} ||| {location}"
            for (contentCity, location) in zip(
                self.contentCity, self.location, strict=True
            )
        ]

        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': specified_location,
            'Timeline': 'N/A',
            'Contacts': self.phoneNumber,
            'Personal Info': self.name,
            'Overall Description': self.description,
            # -----
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-megapersonals-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def capture_screenshot(self, screenshot_name) -> None:
        try:
            self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
            self.create_pdf()
        except Exception as e:
            print(f"Error capturing screenshot: {e}")

    def create_pdf(self) -> None:
        screenshot_files = [os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.description = []
        self.name = []
        self.phoneNumber = []
        self.contentCity = []
        self.location = []
        self.link = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False




