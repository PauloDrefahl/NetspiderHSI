import os
import time
import re
from datetime import datetime
import pandas as pd
from seleniumbase import Driver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from typing_extensions import override
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill


class EscortalligatorScraper(ScraperPrototype):
    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None

        self.cities = [
            "daytona",
            "fort lauderdale",
            "fort myers",
            "gainesville",
            "jacksonville",
            "keys",
            "miami",
            "ocala",
            "okaloosa",
            "orlando",
            "palm bay",
            "panama city",
            "pensacola",
            "bradenton",
            "space coast",
            "st. augustine",
            "tallahassee",
            "tampa",
            "treasure coast",
            "west palm beach",
            "jacksonville"
        ]
        self.city = ''
        self.state = 'florida'
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.keywords: set[str] = set()
        self.flagged_keywords: set[str] = set()
        self.only_posts_with_payment_methods = False
        self.completed = False

        self.join_keywords = False
        self.search_mode = False

        self.keywords_found_in_post: set[str] = set()

        # lists to store data and then send to Excel file
        self.phone_number = []
        self.description = []
        self.links = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.timestamps = []
        self.location = []
        self.age = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return self.cities

    def set_city(self, city) -> None:
        self.city = city.replace(' ', '').replace('.', '')

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords: set[str]) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords: set[str]) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''

    def initialize(self) -> None:
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            # window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()
        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/escortalligator-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/escortalligator-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)

        # Get data from posts
        self.get_data(links)
        self.close_webpage()
        self.reset_variables()
        self.completed = True

    def stop_scraper(self) -> None:
        self.completed = True

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        time.sleep(1)
        self.driver.get(self.url)
        
        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.quit()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> list:
        # Click on the "terms of use" button.
        self.driver.click(By.CLASS_NAME, "button")

        # Click on the "scum warning" button.
        self.driver.click(By.CLASS_NAME, "footer")

        posts = self.driver.find_elements(
            By.CSS_SELECTOR, '#list [href]')
        links = [post.get_attribute('href') for post in posts]
        return links[::3]

    def get_formatted_url(self) -> None:
        if self.city == 'fortlauderdale':
            self.city = 'ftlauderdale'
        self.url = f'https://escortalligator.com.listcrawler.eu/brief/escorts/usa/{self.state}/{self.city}/1'

    def get_data(self, links) -> None:
        links = set(links)
        counter = 1

        for link in links:
            print(f"Processing link {counter}/{len(links)}: {link}")
            try:
                if not self.completed:
                    self.driver.get(link)
                    time.sleep(1)
                    self.driver.get(link)
                    assert "Page not found" not in self.driver.page_source

                    try:
                        timestamp = self.driver.find_element(
                            By.CLASS_NAME, 'postCreatedOn').text
                    except NoSuchElementException:
                        timestamp = 'N/A'

                    try:
                        description = self.driver.find_element(
                            By.CLASS_NAME, 'viewpostbody').text
                    except NoSuchElementException:
                        description = 'N/A'

                    try:
                        phone_number = self.driver.find_element(
                            By.CLASS_NAME, 'userInfoContainer').text
                    except NoSuchElementException:
                        phone_number = 'N/A'

                    try:
                        location_and_age = self.driver.find_element(
                            By.CLASS_NAME, 'viewpostlocationIconBabylon').text
                        location, age = self.parse_location_and_age(location_and_age)
                    except NoSuchElementException:
                        location, age = "N/A", "N/A"

                    # reassign variables for each post
                    self.keywords_found_in_post.clear()

                    # Search the post's contents for keywords.
                    self.check_keywords_found(description, location, age, phone_number, link)

                    if self._should_discard_post(description):
                        continue

                    # Save the data we collected about the post.
                    self.append_data(counter, description, link, location, age, phone_number, timestamp)
                    screenshot_name = str(counter) + ".png"
                    self.capture_screenshot(screenshot_name)
                    counter += 1

                    self.RAW_format_data_to_excel()
                    self.CLEAN_format_data_to_excel()
                # Breaks the links loop for fast closing time once user presses stop scraper
                else:
                    break
            except Exception as e:
                print(f"Error processing link {link}: {e}")
                continue

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, counter, description, link, location, age, phone_number, timestamp) -> None:
        self.post_identifier.append(counter)
        self.phone_number.append(phone_number)
        self.links.append(link)
        self.location.append(location)
        self.age.append(age)
        self.description.append(description)
        payment_methods = self.get_payment_methods(description)
        self.payment_methods_found.append("\n".join(payment_methods) or "N/A")
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(len(self.keywords_found_in_post) or "N/A")
        social_media = self.get_social_media(description)
        self.social_media_found.append("\n".join(social_media) or "N/A")
        self.timestamps.append(timestamp)
        # Store information about the post in the database.
        try:
            with self.open_database() as connection, connection.cursor() as cursor:
                # Escort Alligator displays timestamps in 24-hour notation, but
                # still includes 'AM' and 'PM', which confuses PostgreSQL.
                timestamp = re.sub("AM|PM", "", timestamp)
                cursor.execute(
                    """
                    insert into raw_escort_alligator_posts
                    values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict do nothing;
                    """,
                    (
                        link,
                        self.city,
                        location,
                        timestamp,
                        phone_number,
                        age,
                        description,
                        payment_methods,
                        social_media,
                        list(self.keywords_found_in_post),
                    ),
                )
        except Exception as e:
            print(f"Database write failed: {e}") 

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, description, location, age, phone_number, link):
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(location)
        self.check_and_append_keywords(age)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(link)

    @override
    def check_for_payment_methods(self, description: str) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def get_payment_methods(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        payment_methods: list[str] = []
        for payment_method in self.known_payment_methods:
            if payment_method in description:
                payment_methods.append(payment_method)
        return payment_methods

    def get_social_media(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        social_media: list[str] = []
        for social in self.known_social_media:
            if social in description:
                social_media.append(social)
        return social_media

    @override
    def check_and_append_keywords(self, data: str) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.add(key)

    def _should_discard_post(self, description: str) -> bool:
        if self.join_keywords:
            # Discard posts that don't contain ALL keywords.
            if len(self.keywords_found_in_post) < len(self.keywords):
                return True
        elif not self.only_posts_with_payment_methods and len(self.keywords) > 0:
            # Discard posts that don't contain ANY keywords, unless:
            # 1. We're specifically looking for posts with payment methods, in
            #    which case we keep *all* posts with payment methods.
            # 2. No keywords were originally provided.
            if len(self.keywords_found_in_post) == 0:
                return True

        if self.only_posts_with_payment_methods:
            if not self.check_for_payment_methods(description):
                return True

        return False

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = {
            'Post-identifier': self.post_identifier,
            'Link': self.links,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': self.location,
            # ------
            'Timestamp': self.timestamps,
            # -------
            'Phone-Number': self.phone_number,
            # ---------
            'Age': self.age,
            # ---------
            'Description': self.description,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            # ---------
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        }
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-escortalligator-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(
                    i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        titled_columns = pd.DataFrame({
            # ---- abs identifiers
            'Post-identifier': self.post_identifier,
            'Link': self.links,  # could also be a keyword source too
            # ------- time and place of posting
            'Inputted City / Region': self.city,
            'Specified Location': self.location,  # could also be a keyword source too
            'Timeline': self.timestamps,
            # ------ methods of tracking
            'Contacts': self.phone_number,  # could also be a keyword source too
            # ----- keyword sources
            'Personal Info': self.age,
            'Overall Description': self.description,
            # ------- other forms of transactions and communication
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            # ------- keyword stats
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })

        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-escortalligator-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def parse_location_and_age(self, location_and_age_str):
        # Example input: "Age: 24Location:  Incall and outcall Daytona"
        age = 'N/A'
        location = 'N/A'

        # Splitting the string based on 'Location:' to separate age and location
        parts = location_and_age_str.split('Location:')

        # Extracting age part and removing 'Age:' prefix
        if len(parts) > 0:
            age_part = parts[0].strip()
            age = age_part.replace('Age:', '').strip()

        # Extracting location if present
        if len(parts) > 1:
            location = parts[1].strip()

        return location, age

    def capture_screenshot(self, screenshot_name) -> None:
        try:
            self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
            self.create_pdf()
        except Exception as e:
            print(f"Error capturing screenshot: {e}")

    def create_pdf(self) -> None:
        screenshot_files = [
            os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if
            filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.phone_number = []
        self.description = []
        self.location = []
        self.age = []
        self.links = []
        self.post_identifier = []
        self.timestamps = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.social_media_found = []
