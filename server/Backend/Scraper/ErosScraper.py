from Backend.ScraperPrototype import ScraperPrototype
from datetime import datetime
from seleniumbase import Driver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
import pandas as pd
import os
import time
import img2pdf
from openpyxl.styles import PatternFill
from typing_extensions import override


class ErosScraper(ScraperPrototype):
    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None

        self.cities = {
            "miami": 'https://www.eros.com/florida/miami/sections/miami_escorts.htm',
            "naples": 'https://www.eros.com/florida/naples/sections/naples_escorts.htm',
            "north florida": 'https://www.eros.com/florida/north_florida/sections/north_florida_escorts.htm',
            "orlando": 'https://www.eros.com/florida/tampa/sections/tampa_escorts.htm',
            "tampa": 'https://www.eros.com/florida/tampa/sections/tampa_escorts.htm'
        }
        self.city = ''
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.keywords: set[str] = set()
        self.flagged_keywords: set[str] = set()
        self.completed = False

        self.join_keywords = False
        self.search_mode = False

        self.keywords_found_in_post: set[str] = set()
        self.only_posts_with_payment_methods = False
        self.social_media_found = []

        # lists to store data and then send to Excel file
        self.post_identifier = []
        self.link = []
        self.profile_header = []
        self.about_info = []
        self.info_details = []
        self.contact_details = []
        self.payment_methods_found = []

        self.number_of_keywords_found = []
        self.keywords_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords: set[str]) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords: set[str]) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''
    def initialize(self) -> None:
        # Date and time of search
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            # window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()

        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/eros-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/eros-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)

        # Get data from posts
        self.get_data(links)
        self.close_webpage()
        self.reset_variables()
        self.completed = True

    def stop_scraper(self) -> None:
        self.completed = True

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        time.sleep(1)
        self.driver.get(self.url)
        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.quit()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> set:
        try:
            # Find website agreement
            self.driver.find_element(
                By.XPATH, '//*[@id="agree_enter_website"]').click()
            self.driver.find_element(
                By.XPATH, '// *[ @ id = "ageModal"] / div / div / div[2] / button').click()
        except NoSuchElementException:
            exit(1)

        try:
            # Find all profile links
            posts = self.driver.find_elements(
                By.CSS_SELECTOR, '#listing > div.grid.fourPerRow.mobile.switchable [href]')
        except NoSuchElementException:
            exit(1)

        if posts:
            links = [post.get_attribute('href') for post in posts]
        else:
            exit(1)

        return set(links)

    def get_formatted_url(self) -> None:
        self.url = self.cities.get(self.city)

    def get_data(self, links) -> None:
        counter = 1

        for link in links:
            print(f"Processing link {counter}/{len(links)}: {link}")
            if not self.completed:
                self.driver.implicitly_wait(10)
                self.driver.get(link)
                time.sleep(1)
                self.driver.get(link)
                assert "Page not found" not in self.driver.page_source

                try:
                    profile_header = self.driver.find_element(
                        By.XPATH, '//*[@id="pageone"]/div[1]').text
                except NoSuchElementException:
                    profile_header = 'N/A'

                try:
                    description = self.driver.find_element(
                        By.XPATH, '// *[ @ id = "pageone"] / div[3] / div / div[1] / div[2]').text
                except NoSuchElementException:
                    description = 'N/A'

                try:
                    info_details = self.driver.find_element(
                        By.XPATH, '//*[@id="pageone"]/div[3]/div/div[2]/div[1]/div').text
                except NoSuchElementException:
                    info_details = 'N/A'

                try:
                    contact_details = self.driver.find_element(
                        By.XPATH, '//*[@id="pageone"]/div[3]/div/div[2]/div[2]').text
                except NoSuchElementException:
                    contact_details = 'N/A'

                # reassign variables for each post
                self.keywords_found_in_post.clear()

                # Search the post's contents for keywords.
                self.check_keywords_found(contact_details, description, info_details, profile_header, link)

                if self._should_discard_post(description):
                    continue

                # Save the data we collected about the post.
                self.append_data(contact_details, counter, description, info_details, link, profile_header)
                screenshot_name = str(counter) + ".png"
                self.capture_screenshot(screenshot_name)
                counter += 1

                self.RAW_format_data_to_excel()
                self.CLEAN_format_data_to_excel()
            # Breaks the links loop for fast closing time once user presses stop scraper
            else:
                break

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, contact_details, counter, description, info_details, link, profile_header) -> None:
        self.post_identifier.append(counter)
        self.link.append(link)
        self.profile_header.append(profile_header)
        self.about_info.append(description)
        self.info_details.append(info_details)
        self.contact_details.append(contact_details)
        payment_methods = self.get_payment_methods(description)
        self.payment_methods_found.append("\n".join(payment_methods) or "N/A")
        social_media = self.get_social_media(description)
        self.social_media_found.append("\n".join(social_media) or "N/A")
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(len(self.keywords_found_in_post) or "N/A")
        # Store information about the post in the database.
        try:
            with self.open_database() as connection, connection.cursor() as cursor:
                cursor.execute(
                    """
                    insert into raw_eros_posts
                    values (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict do nothing;
                    """,
                    (
                        link,
                        self.city,
                        profile_header,
                        description,
                        info_details,
                        contact_details,
                        payment_methods,
                        social_media,
                        list(self.keywords_found_in_post),
                    ),
                )
        except Exception as e:
            print(f"Database write failed: {e}") 

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, contact_details, description, info_details, profile_header, link):
        self.check_and_append_keywords(contact_details)
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(info_details)
        self.check_and_append_keywords(profile_header)
        self.check_and_append_keywords(link)

    @override
    def check_for_payment_methods(self, description: str) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def get_payment_methods(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        payment_methods: list[str] = []
        for payment_method in self.known_payment_methods:
            if payment_method in description:
                payment_methods.append(payment_method)
        return payment_methods

    def get_social_media(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        social_media: list[str] = []
        for social in self.known_social_media:
            if social in description:
                social_media.append(social)
        return social_media

    @override
    def check_and_append_keywords(self, data: str) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.add(key)

    def _should_discard_post(self, description: str) -> bool:
        if self.join_keywords:
            # Discard posts that don't contain ALL keywords.
            if len(self.keywords_found_in_post) < len(self.keywords):
                return True
        elif not self.only_posts_with_payment_methods and len(self.keywords) > 0:
            # Discard posts that don't contain ANY keywords, unless:
            # 1. We're specifically looking for posts with payment methods, in
            #    which case we keep *all* posts with payment methods.
            # 2. No keywords were originally provided.
            if len(self.keywords_found_in_post) == 0:
                return True

        if self.only_posts_with_payment_methods:
            if not self.check_for_payment_methods(description):
                return True

        return False

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = {
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            'Inputted City / Region': self.city,
            'Specified Location': 'N/A',
            # -------
            'Profile-header': self.profile_header,
            'About-info': self.about_info,
            'Info-details': self.info_details,
            'Contact-details': self.contact_details,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        }

        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-eros-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']

            for i in range(2, worksheet.max_row):
                keywords = worksheet["H" + str(
                    i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["H" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        # Concatenate attributes to fit into the CLEAN spreadsheet format, which
        # is consistent across all scrapers.
        overall_description = [
            f"{profile_header} ||| {info_details}"
            for (profile_header, info_details) in zip(
                self.profile_header, self.info_details, strict=True
            )
        ]

        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': 'N/A',
            'Timeline': 'N/A',
            'Contacts': self.contact_details,
            'Personal Info': self.about_info,
            'Overall Description': overall_description,
            # ------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-eros-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width


    def capture_screenshot(self, screenshot_name) -> None:
        try:
            self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
            self.create_pdf()
        except Exception as e:
            print(f"Error capturing screenshot: {e}")

    def create_pdf(self) -> None:
        screenshot_files = [
            os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.post_identifier = []
        self.link = []
        self.profile_header = []
        self.about_info = []
        self.info_details = []
        self.contact_details = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False




