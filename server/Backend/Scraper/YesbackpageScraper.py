import os
import time
from collections import defaultdict
from datetime import datetime
import pandas as pd
from seleniumbase import Driver
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from typing_extensions import override
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill
import threading


class YesbackpageScraper(ScraperPrototype):

    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None
        self.cities = {
            "florida": 'https://www.yesbackpage.com/-10/posts/8-Adult/',
            "broward": 'https://www.yesbackpage.com/70/posts/8-Adult/',
            "daytona beach": 'https://www.yesbackpage.com/71/posts/8-Adult/',
            "florida keys": 'https://www.yesbackpage.com/67/posts/8-Adult/',
            "ft myers-sw florida": 'https://www.yesbackpage.com/68/posts/8-Adult/',
            "gainesville": 'https://www.yesbackpage.com/72/posts/8-Adult/',
            "jacksonville": 'https://www.yesbackpage.com/73/posts/8-Adult/',
            "lakeland": 'https://www.yesbackpage.com/74/posts/8-Adult/',
            "miami": 'https://www.yesbackpage.com/75/posts/8-Adult/',
            "ocala": 'https://www.yesbackpage.com/76/posts/8-Adult/',
            "orlando": 'https://www.yesbackpage.com/77/posts/8-Adult/',
            "palm beach": 'https://www.yesbackpage.com/69/posts/8-Adult/',
            "panama city": 'https://www.yesbackpage.com/78/posts/8-Adult/',
            "pensacola-panhandle": 'https://www.yesbackpage.com/79/posts/8-Adult/',
            "sarasota-brandenton": 'https://www.yesbackpage.com/80/posts/8-Adult/',
            "space coast": 'https://www.yesbackpage.com/81/posts/8-Adult/',
            "st augustine": 'https://www.yesbackpage.com/82/posts/8-Adult/',
            "tallahassee": 'https://www.yesbackpage.com/83/posts/8-Adult/',
            "tampa bay area": 'https://www.yesbackpage.com/84/posts/8-Adult/',
            "treasure coast": 'https://www.yesbackpage.com/85/posts/8-Adult/',
            "west palm beach": 'https://www.yesbackpage.com/679/posts/8-Adult/'
        }
        self.city = ''
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.pdf = None
        self.keywords: set[str] = set()
        self.flagged_keywords: set[str] = set()
        self.search_mode = False
        self.completed = False

        self.join_keywords = False
        self.keywords_found_in_post: set[str] = set()

        self.only_posts_with_payment_methods = False

        # lists to store data and then send to csv file

        self.phone_number = []
        self.posted_on = []
        self.expires_on = []
        self.reply_to = []
        self.link = []
        self.name = []
        self.sex = []
        self.email = []
        self.location = []
        self.description = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.services = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords: set[str]) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords: set[str]) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''
    def initialize(self) -> None:
        # set up directories to save screenshots and Excel file.
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        self.driver = Driver(
            # Download the latest ChromeDriver for the current major version.
            driver_version="mlatest",
            undetectable=True,
            uc_subprocess=True,
            headless=self.search_mode,
            # Override the default mode (headless mode) on Linux.
            headed=not self.search_mode,
            # Use a fixed window size in headless mode.
            # window_size="1920,1080" if self.search_mode else None,
        )

        # Open Webpage with URL
        self.open_webpage()

        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/yesbackpage-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/yesbackpage-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)
        self.get_data(links)
        self.close_webpage()
        self.reset_variables()
        self.completed = True

    def stop_scraper(self) -> None:
        self.completed = True

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        # NOTE: Maximizing the window in headless mode makes it too big:
        # https://chromium.googlesource.com/chromium/src.git/+/f2bdeab65/ui/views/win/hwnd_message_handler_headless.cc#264
        if not self.search_mode:
            self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.quit()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> list:
        posts = self.driver.find_elements(
            By.CLASS_NAME, 'posttitle')
        links = [post.get_attribute('href') for post in posts]
        return links[2:]

    def get_formatted_url(self) -> None:
        self.url = self.cities.get(self.city)

    def get_data(self, links) -> None:
        counter = 1

        for link in links:
            print(f"Processing link {counter}/{len(links)}: {link}")
            if not self.completed:
                self.driver.implicitly_wait(10)
                self.driver.get(link)
                time.sleep(1)
                self.driver.get(link)
                assert "Page not found" not in self.driver.page_source

                try:
                    description = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/table[2]/tbody/'
                                'tr/td/div/p[2]').text
                except NoSuchElementException:
                    description = 'N/A'

                try:
                    timestamp = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/table[1]/tbody/'
                                'tr/td/div[3]/div[1]').text
                    posted_on, expires_on, reply_to = self.parse_timestamp(timestamp)
                except NoSuchElementException:
                    posted_on = 'N/A'
                    expires_on = 'N/A'
                    reply_to = 'N/A'

                # check if page contains "col-sm-6 offset-sm-3" which is the table that contains name, phone number, etc.
                if self.driver.find_elements(By.XPATH, '//*[@id="mainCellWrapper"]/div[1]/table/tbody/tr[1]/td/div[1]/div'):
                    try:
                        name = self.driver.find_element(
                            By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                    'tbody/tr[1]/td[2]').text[2:]
                    except NoSuchElementException:
                        name = 'N/A'

                    try:
                        sex = self.driver.find_element(
                            By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                    'tbody/tr[2]/td[2]').text[2:]
                    except NoSuchElementException:
                        sex = 'N/A'

                    try:
                        phone_number = self.driver.find_element(
                            By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                    'tbody/tr[6]/td[2]').text[2:]
                    except NoSuchElementException:
                        phone_number = 'NA'

                    try:
                        email = self.driver.find_element(
                            By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                    'tbody/tr[8]/td[2]').text[2:]
                        # email = self.validate_email(email)
                    except NoSuchElementException:
                        email = 'N/A'

                    try:
                        location = self.driver.find_element(
                            By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                    'tbody/tr[9]/td[2]').text[2:]
                    except NoSuchElementException:
                        location = 'N/A'

                    try:
                        services = self.driver.find_element(
                            By.XPATH, '//*[@id="mainCellWrapper"]/div/table/tbody/tr/td/div[1]/div/table/'
                                    'tbody/tr[5]/td[2]').text[2:]
                    except NoSuchElementException:
                        services = 'N/A'
                else:
                    posted_on = 'N/A'
                    expires_on = 'N/A'
                    reply_to = 'N/A'
                    name = 'N/A'
                    sex = 'N/A'
                    phone_number = 'N/A'
                    email = 'N/A'
                    location = 'N/A'
                    services = 'N/A'

                # reassign variables for each post
                self.keywords_found_in_post.clear()

                # Search the post's contents for keywords.
                self.check_keywords_found(description, name, sex, phone_number, email, location, services, link)

                if self._should_discard_post(description):
                    continue

                # Save the data we collected about the post.
                self.append_data(counter, description, email, link, location, name, phone_number, services, sex, posted_on, expires_on, reply_to)
                screenshot_name = str(counter) + ".png"
                self.capture_screenshot(screenshot_name)
                counter += 1

                self.RAW_format_data_to_excel()
                self.CLEAN_format_data_to_excel()
            # Breaks the links loop for fast closing time once user presses stop scraper
            else:
                break

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, counter, description, email, link, location, name, phone_number, services, sex, posted_on, expires_on, reply_to):
        self.post_identifier.append(counter)
        self.posted_on.append(posted_on)
        self.expires_on.append(expires_on)
        self.reply_to.append(reply_to)
        self.link.append(link)
        self.description.append(description)
        self.name.append(name)
        self.sex.append(sex)
        self.phone_number.append(phone_number)
        self.email.append(email)
        self.location.append(location)
        payment_methods = self.get_payment_methods(description)
        self.payment_methods_found.append("\n".join(payment_methods) or "N/A")
        self.services.append(services)
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(len(self.keywords_found_in_post) or "N/A")
        social_media = self.get_social_media(description)
        self.social_media_found.append("\n".join(social_media) or "N/A")
        # Store information about the post in the database.
        try:
            with self.open_database() as connection, connection.cursor() as cursor:
                # Map strings to their corresponding enum labels.
                sex_enum_map = defaultdict(
                    lambda: "Other", {"Male": "Male", "Female": "Female"}
                )
                # Convert missing timestamps to `None`.
                posted_on = None if posted_on == "N/A" else posted_on
                expires_on = None if expires_on == "N/A" else expires_on
                cursor.execute(
                    """
                    insert into raw_yesbackpage_posts
                    values (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    on conflict do nothing;
                    """,
                    (
                        link,
                        self.city,
                        location,
                        posted_on,
                        expires_on,
                        phone_number,
                        email,
                        name,
                        sex_enum_map[sex],
                        reply_to,
                        description,
                        services,
                        payment_methods,
                        social_media,
                        list(self.keywords_found_in_post),
                    ),
                )
        except Exception as e:
            print(f"Database write failed: {e}") 

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, description, name, sex, phone_number, email, location, services, link) -> None:
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(name)
        self.check_and_append_keywords(sex)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(email)
        self.check_and_append_keywords(location)
        self.check_and_append_keywords(services)
        self.check_and_append_keywords(link)

    @override
    def check_for_payment_methods(self, description: str) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def get_payment_methods(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        payment_methods: list[str] = []
        for payment_method in self.known_payment_methods:
            if payment_method in description:
                payment_methods.append(payment_method)
        return payment_methods

    def get_social_media(self, description: str) -> list[str]:
        # Normalize the case of the description.
        description = description.lower()
        social_media: list[str] = []
        for social in self.known_social_media:
            if social in description:
                social_media.append(social)
        return social_media

    @override
    def check_and_append_keywords(self, data: str) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.add(key)

    def _should_discard_post(self, description: str) -> bool:
        if self.join_keywords:
            # Discard posts that don't contain ALL keywords.
            if len(self.keywords_found_in_post) < len(self.keywords):
                return True
        elif not self.only_posts_with_payment_methods and len(self.keywords) > 0:
            # Discard posts that don't contain ANY keywords, unless:
            # 1. We're specifically looking for posts with payment methods, in
            #    which case we keep *all* posts with payment methods.
            # 2. No keywords were originally provided.
            if len(self.keywords_found_in_post) == 0:
                return True

        if self.only_posts_with_payment_methods:
            if not self.check_for_payment_methods(description):
                return True

        return False

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': self.location,
            'Posted On': self.posted_on,
            'Expires On': self.expires_on,
            'Phone-Number': self.phone_number,
            'E-mail': self.email,
            'Name': self.name,
            'Sex': self.sex,
            'Reply To': self.reply_to,
            'Description': self.description,
            'Services': self.services,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-yesbackpage-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["N" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["N" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def CLEAN_format_data_to_excel(self) -> None:
        # Concatenate attributes to fit into the CLEAN spreadsheet format, which
        # is consistent across all scrapers.
        personal_info = [
            f"{name} ||| {sex}"
            for (name, sex) in zip(self.name, self.sex, strict=True)
        ]

        contacts = [
            f"{phone_number} ||| {email}"
            for (phone_number, email) in zip(
                self.phone_number, self.email, strict=True
            )
        ]

        overall_description = [
            f"{description} ||| {services} ||| {reply_to}"
            for (description, services, reply_to) in zip(
                self.description, self.services, self.reply_to, strict=True
            )
        ]

        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Link': self.link,
            # ------
            'Inputted City / Region': self.city,
            'Specified Location': self.location,
            'Timeline': self.posted_on,
            'Contacts': contacts,
            'Personal Info': personal_info,
            'Overall Description': overall_description,
            # -----
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-yesbackpage-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def parse_timestamp(self, timestamp_str):
        # Initialize all parts as 'N/A' to handle missing parts
        posted_on, updated_on, expires_on, reply_to = ('N/A',) * 4

        # Split the string by 'Expires On:' to separate the first half and the 'Expires On:' part
        parts = timestamp_str.split('Expires On:')
        first_half = parts[0].strip() if len(parts) > 0 else 'N/A'

        # Check if 'Updated On:' is in the first half
        if 'Updated On:' in first_half:
            posted_on, updated_on = first_half.split('Updated On:')
            posted_on = posted_on.replace('Posted on:', '').strip()
            updated_on = updated_on.strip()
        else:
            # If there's no 'Updated On:', the whole first half is 'Posted on:'
            posted_on = first_half.replace('Posted on:', '').strip()

        # Now, handle the 'Expires On:' part
        if len(parts) > 1:
            second_half = parts[1].strip()
            if 'Reply to:' in second_half:
                expires_on, reply_to = second_half.split('Reply to:')
                expires_on = expires_on.strip()
                reply_to = reply_to.strip()
            else:
                expires_on = second_half.strip()

        return posted_on, expires_on, reply_to

    def validate_email(self, email):
        return email if "@" in email else "N/A"

    def capture_screenshot(self, screenshot_name) -> None:
        try:
            self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
            self.create_pdf()
        except Exception as e:
            print(f"Error capturing screenshot: {e}")

    def create_pdf(self) -> None:
        screenshot_files = [os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.phone_number = []
        self.posted_on = []
        self.expires_on = []
        self.reply_to = []
        self.link = []
        self.name = []
        self.sex = []
        self.email = []
        self.location = []
        self.description = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.services = []
        self.number_of_keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.keywords_found = []
        self.social_media_found = []
