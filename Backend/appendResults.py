import pandas as pd
import os
from datetime import datetime
from shutil import copy2
from PyPDF2 import PdfMerger
import openpyxl
from openpyxl.utils import get_column_letter
pdf_merger = PdfMerger()

'''
    ---------------------------------
    Inputs for files to append and directory to save
    ---------------------------------
'''

# Define the main results directory
results_directory = 'C:\\Users\\kskos\\PycharmProjects\\HSI_Back_Test3\\result'

# List of selected folders to process
selected_folders = [  # make this into a json with file names
    'megapersonals-daytona-2024-02-21_00-57-38',
    'yesbackpage-florida-2024-02-21_00-59-55'
]

'''
    ---------------------------------
    Making the directories and paths
    ---------------------------------
'''

# Generate the new folder name by combining selected folder names and current date and time
new_folder_name = '-'.join(folder.split('-')[0] for folder in selected_folders)
new_folder_name += '-' + '-'.join(folder.split('-')[1] for folder in selected_folders)
new_folder_name += '-' + datetime.now().strftime('%Y-%m-%d_%H-%M-%S')

# Create the new folder in the results directory
new_folder_path = os.path.join(results_directory, new_folder_name)
os.makedirs(new_folder_path, exist_ok=True)

# ----

# Create the new screenshot folder in the new folder directory
new_screenshot_folder_path = os.path.join(new_folder_path, 'screenshots')
os.makedirs(new_screenshot_folder_path, exist_ok=True)

# ----

# The new spreadsheet file has the same name as the new folder with a .xlsx extension
new_file_name = 'CLEAN-' + new_folder_name
new_file_path = os.path.join(new_folder_path, new_file_name + '.xlsx')

'''
    ---------------------------------
    Saving Process
    ---------------------------------
'''

# Initialize an empty list to store the DataFrames
dataframes = []
folder_count = 0
# Loop through the selected folders
for folder in selected_folders:

    folder_path = os.path.join(results_directory, folder)

    # The spreadsheet file has the same name as the folder with a .xlsx extension
    spreadsheet_name = 'CLEAN-' + folder + '.xlsx'
    file_path = os.path.join(folder_path, spreadsheet_name)
    if os.path.exists(file_path):
        df = pd.read_excel(file_path)
        dataframes.append(df)

# -----

    # Copy PNG files from 'screenshots' subfolder
    screenshots_path = os.path.join(folder_path, 'screenshots')
    if os.path.exists(screenshots_path):
        for i, file in enumerate(os.listdir(screenshots_path)):
            if file.endswith('.png'):
                new_screenshot_file_name = f"{folder_count}_{i}.png"
                source_file_path = os.path.join(screenshots_path, file)
                destination_file_path = os.path.join(new_screenshot_folder_path, new_screenshot_file_name)
                print(new_screenshot_folder_path)
                print(destination_file_path)
                copy2(source_file_path, destination_file_path)

    folder_count += 1
    # -----

    # Add PDFs to the merger
    pdf_path = os.path.join(folder_path, 'screenshots')
    if os.path.exists(pdf_path):
        for file in os.listdir(pdf_path):
            if file.endswith('.pdf'):
                pdf_merger.append(os.path.join(pdf_path, file))

'''
    ---------------------------------
    Store the saved information to files and folders
    ---------------------------------
'''
# Concatenate all DataFrames in the list
combined_df = pd.concat(dataframes, ignore_index=True)

# If you want to make sure that no NaN values remain, do another fill na here
combined_df.fillna('N/A', inplace=True)

# Reset the 'Post-identifier' to be a continuous sequence of numbers
#combined_df['Continuous_Post_identifier'] = range(len(combined_df))

# Replace the existing 'Post-identifier' column with a continuous sequence of numbers
combined_df['Post-identifier'] = range(len(combined_df))

# Save the appended DataFrame to the new spreadsheet inside the new folder
combined_df.to_excel(new_file_path, index=False)
workbook = openpyxl.load_workbook(new_file_path)
worksheet = workbook.active

# Loop through each column in the worksheet
for col in worksheet.columns:
    max_length = 0
    column = col[0].column  # Get the column name
    # Iterate over all cells in a column to find the maximum length
    for cell in col:
        try:  # Necessary to avoid error on empty cells
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        except:
            pass
    # Adjust the column width
    adjusted_width = (max_length + 2)
    worksheet.column_dimensions[get_column_letter(column)].width = adjusted_width

# Save the workbook
workbook.save(new_file_path)
workbook.close()

# -----

# Merge and save the PDF files into the new folder
merged_pdf_path = os.path.join(new_screenshot_folder_path, new_folder_name + '.pdf')
pdf_merger.write(merged_pdf_path)
pdf_merger.close()


