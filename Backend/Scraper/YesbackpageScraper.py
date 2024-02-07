import os
import time
from datetime import datetime
import pandas as pd
import undetected_chromedriver as uc
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from Backend.ScraperPrototype import ScraperPrototype
import img2pdf
from openpyxl.styles import PatternFill
import threading


class YesbackpageScraper(ScraperPrototype):

    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None
        self.cities = {
            "florida": 'https://www.yesbackpage.com/-10/posts/8-Adult/',
            "broward": 'https://www.yesbackpage.com/70/posts/8-Adult/',
            "daytona beach": 'https://www.yesbackpage.com/71/posts/8-Adult/',
            "florida keys": 'https://www.yesbackpage.com/67/posts/8-Adult/',
            "ft myers-sw florida": 'https://www.yesbackpage.com/68/posts/8-Adult/',
            "gainesville": 'https://www.yesbackpage.com/72/posts/8-Adult/',
            "jacksonville": 'https://www.yesbackpage.com/73/posts/8-Adult/',
            "lakeland": 'https://www.yesbackpage.com/74/posts/8-Adult/',
            "miami": 'https://www.yesbackpage.com/75/posts/8-Adult/',
            "ocala": 'https://www.yesbackpage.com/76/posts/8-Adult/',
            "orlando": 'https://www.yesbackpage.com/77/posts/8-Adult/',
            "palm beach": 'https://www.yesbackpage.com/69/posts/8-Adult/',
            "panama city": 'https://www.yesbackpage.com/78/posts/8-Adult/',
            "pensacola-panhandle": 'https://www.yesbackpage.com/79/posts/8-Adult/',
            "sarasota-brandenton": 'https://www.yesbackpage.com/80/posts/8-Adult/',
            "space coast": 'https://www.yesbackpage.com/81/posts/8-Adult/',
            "st augustine": 'https://www.yesbackpage.com/82/posts/8-Adult/',
            "tallahassee": 'https://www.yesbackpage.com/83/posts/8-Adult/',
            "tampa bay area": 'https://www.yesbackpage.com/84/posts/8-Adult/',
            "treasure coast": 'https://www.yesbackpage.com/85/posts/8-Adult/',
            "west palm beach": 'https://www.yesbackpage.com/679/posts/8-Adult/'
        }
        self.city = ''
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.pdf = None
        self.keywords = set()
        self.flagged_keywords = set()
        self.search_mode = False
        self.completed = False

        self.join_keywords = False
        self.number_of_keywords_in_post = 0
        self.keywords_found_in_post = []

        self.only_posts_with_payment_methods = False

        # lists to store data and then send to csv file

        self.phone_number = []
        self.timestamp = []
        self.link = []
        self.name = []
        self.sex = []
        self.email = []
        self.location = []
        self.description = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.services = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    def get_cities(self) -> list:
        return list(self.cities.keys())

    def set_city(self, city) -> None:
        self.city = city

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords) -> None:
        self.keywords = keywords

    def initialize(self) -> None:
        # set keywords value
        # self.keywords = keywords
        # set up directories to save screenshots and Excel file.
        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()
        # self.set_search_mode(self.search_mode)
        options = uc.ChromeOptions()
        # TODO - uncomment this line to run headless
        # options.add_argument('--headless')
        options.headless = self.search_mode  # This determines if you program runs headless or not
        self.driver = uc.Chrome(subprocress=True, options=options)

        # Open Webpage with URL
        self.open_webpage()

        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/yesbackpage_{self.city}_{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/yesbackpage.pdf'
        os.mkdir(self.screenshot_directory)
        print("number of threads while running: ", threading.active_count())
        print("keywords inside scraper:", self.keywords)
        self.get_data(links)
        print("get data done")
        # time.sleep(5)
        self.stop_scraper()
        print("closed webpage")
        self.reset_variables()
        print("reset variables")
        self.completed = True
        print("done scraping")

    def stop_scraper(self) -> None:
        self.driver.close()
        self.driver.quit()

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        if self.search_mode:
            self.driver.get(self.url)
        else:
            self.driver.execute_script(f'window.open("{self.url}", "_blank");')
            original_window = self.driver.current_window_handle
            time.sleep(5)
            if len(self.driver.window_handles) >= 2:
                # Switch to the new tab
                self.driver.switch_to.window(self.driver.window_handles[-1])

                # Close the original tab
                self.driver.switch_to.window(original_window)
                self.driver.close()

                # Switch back to the new tab
                self.driver.switch_to.window(self.driver.window_handles[-1])
        self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.close()

    def get_links(self) -> list:
        posts = self.driver.find_elements(
            By.CLASS_NAME, 'posttitle')
        links = [post.get_attribute('href') for post in posts]
        return links[2:]

    def get_formatted_url(self) -> None:
        self.url = self.cities.get(self.city)

    def get_data(self, links) -> None:
        links = links

        counter = 0

        for link in links:
            self.driver.implicitly_wait(10)
            self.driver.get(link)
            assert "Page not found" not in self.driver.page_source

            try:
                description = self.driver.find_element(
                    By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/table[2]/tbody/'
                              'tr/td/div/p[2]').text
            except NoSuchElementException:
                description = 'N/A'

            try:
                timestamp = self.driver.find_element(
                    By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/table[1]/tbody/'
                              'tr/td/div[3]/div[1]').text
            except NoSuchElementException:
                timestamp = 'Not exist'

            # check if page contains "col-sm-6 offset-sm-3" which is the table that contains name, phone number, etc.
            if self.driver.find_elements(By.XPATH, '//*[@id="mainCellWrapper"]/div[1]/table/tbody/tr[1]/td/div[1]/div'):
                try:
                    name = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                  'tbody/tr[1]/td[2]').text[2:]
                except NoSuchElementException:
                    name = 'N/A'

                try:
                    sex = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                  'tbody/tr[2]/td[2]').text[2:]
                except NoSuchElementException:
                    sex = 'N/A'

                try:
                    phone_number = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                  'tbody/tr[6]/td[2]').text[2:]
                except NoSuchElementException:
                    phone_number = 'NA'

                try:
                    email = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                  'tbody/tr[8]/td[2]').text[2:]
                except NoSuchElementException:
                    email = 'N/A'

                try:
                    location = self.driver.find_element(
                        By.XPATH, '/html/body/div[3]/div/div[1]/table/tbody/tr[1]/td/div[1]/div/table/'
                                  'tbody/tr[9]/td[2]').text[2:]
                except NoSuchElementException:
                    location = 'N/A'

                try:
                    services = self.driver.find_element(
                        By.XPATH, '//*[@id="mainCellWrapper"]/div/table/tbody/tr/td/div[1]/div/table/'
                                  'tbody/tr[5]/td[2]').text[2:]
                except NoSuchElementException:
                    services = 'N/A'
            else:
                timestamp = 'not exist 2'
                name = 'N/A'
                sex = 'N/A'
                phone_number = 'N/A'
                email = 'N/A'
                location = 'N/A'
                services = 'N/A'

            # reassign variables for each post
            self.number_of_keywords_in_post = 0
            self.keywords_found_in_post = []

            if self.join_keywords and self.only_posts_with_payment_methods:
                if self.check_keywords(description) or self.check_keywords(name) or self.check_keywords(sex) \
                        or self.check_keywords(phone_number) or self.check_keywords(email) \
                        or self.check_keywords(location) or self.check_keywords(services):
                    self.check_keywords_found(description, name, sex, phone_number, email, location, services)
                    counter = self.join_with_payment_methods(counter, description, email, link, location, name,
                                                             phone_number, services, sex, timestamp)

            elif self.join_keywords or self.only_posts_with_payment_methods:
                if self.join_keywords:
                    if self.check_keywords(description) or self.check_keywords(name) or self.check_keywords(sex) \
                            or self.check_keywords(phone_number) or self.check_keywords(email) \
                            or self.check_keywords(location) or self.check_keywords(services):
                        self.check_keywords_found(description, name, sex, phone_number, email, location, services)
                        counter = self.join_inclusive(counter, description, email, link, location, name, phone_number,
                                                      services, sex, timestamp)

                elif self.only_posts_with_payment_methods:
                    if len(self.keywords) > 0:
                        if self.check_keywords(description) or self.check_keywords(name) or self.check_keywords(sex) \
                                or self.check_keywords(phone_number) or self.check_keywords(email) \
                                or self.check_keywords(location) or self.check_keywords(services):
                            self.check_keywords_found(description, name, sex, phone_number, email, location, services)

                    counter = self.payment_methods_only(counter, description, email, link, location, name,
                                                        phone_number, services, sex, timestamp)
            else:
                # run if keywords
                if len(self.keywords) > 0:
                    if self.check_keywords(description) or self.check_keywords(name) or self.check_keywords(sex) \
                            or self.check_keywords(phone_number) or self.check_keywords(email) \
                            or self.check_keywords(location) or self.check_keywords(services):
                        self.check_keywords_found(description, name, sex, phone_number, email, location, services)
                        self.append_data(counter, description, email, link, location, name, phone_number, services,
                                         sex, timestamp)
                        screenshot_name = str(counter) + ".png"
                        self.capture_screenshot(screenshot_name)
                        counter += 1
                else:
                    self.append_data(counter, description, email, link, location, name, phone_number, services,
                                     sex, timestamp)
                    screenshot_name = str(counter) + ".png"
                    self.capture_screenshot(screenshot_name)
                    counter += 1
            self.format_data_to_excel()

    def join_with_payment_methods(self, counter, description, email, link, location, name, phone_number,
                                  services, sex, timestamp) -> int:
        if self.check_for_payment_methods(description) and len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(counter, description, email, link, location, name, phone_number, services,
                             sex, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def check_keywords_found(self, description, name, sex, phone_number, email, location, services) -> None:
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(name)
        self.check_and_append_keywords(sex)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(email)
        self.check_and_append_keywords(location)
        self.check_and_append_keywords(services)

    def append_data(self, counter, description, email, link, location, name, phone_number, services, sex, timestamp):
        self.post_identifier.append(counter)
        self.timestamp.append(timestamp)
        self.link.append(link)
        self.description.append(description)
        self.name.append(name)
        self.sex.append(sex)
        self.phone_number.append(phone_number)
        self.email.append(email)
        self.location.append(location)
        self.check_and_append_payment_methods(description)
        self.services.append(services)
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(self.number_of_keywords_in_post or 'N/A')
        self.check_for_social_media(description)

    def format_data_to_excel(self) -> None:
        titled_columns = pd.DataFrame({
            'Post-identifier': self.post_identifier,
            'Timestamp': self.timestamp,
            'Phone-Number': self.phone_number,
            'Link': self.link,
            'Location': self.location,
            'Name': self.name,
            'Sex': self.sex,
            'E-mail': self.email,
            'Services': self.services,
            'Description': self.description,
            'payment-methods': self.payment_methods_found,
            'keywords-found': self.keywords_found,
            'number-of-keywords-found': self.number_of_keywords_found,
            'social-media-found': self.social_media_found
        })
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/yesbackpage-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["L" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["L" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def reset_variables(self) -> None:
        self.phone_number = []
        self.timestamp = []
        self.link = []
        self.name = []
        self.sex = []
        self.email = []
        self.location = []
        self.description = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.services = []
        self.number_of_keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.keywords_found = []
        self.social_media_found = []

    def check_for_payment_methods(self, description) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def check_and_append_payment_methods(self, description) -> None:
        payments = ''
        for payment in self.known_payment_methods:
            if payment in description.lower():
                payments += payment + '\n'

        if payments != '':
            self.payment_methods_found.append(payments)
        else:
            self.payment_methods_found.append('N/A')

    def check_for_social_media(self, description) -> None:
        social_media = ''
        for social in self.known_social_media:
            if social in description.lower():
                social_media += social + '\n'

        if social_media != '':
            self.social_media_found.append(social_media)
        else:
            self.social_media_found.append('N/A')

    def capture_screenshot(self, screenshot_name) -> None:
        self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
        self.create_pdf()

    def create_pdf(self) -> None:
        screenshot_files = [os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def check_keywords(self, data) -> bool:
        for key in self.keywords:
            if key in data.lower():
                return True
        return False

    def check_and_append_keywords(self, data) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.append(key)
                self.number_of_keywords_in_post += 1

    def join_inclusive(self, counter, description, email, link, location, name, phone_number, services, sex, timestamp) -> int:
        if len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(counter, description, email, link, location, name, phone_number, services,
                             sex, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def payment_methods_only(self, counter, description, email, link, location, name, phone_number,
                             services, sex, timestamp) -> int:

        if self.check_for_payment_methods(description):
            self.append_data(counter, description, email, link, location, name, phone_number, services,
                             sex, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter
