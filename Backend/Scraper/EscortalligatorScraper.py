import os
import re
import time
from datetime import datetime
import pandas as pd
from selenium.common import NoSuchElementException
from selenium.webdriver.common.by import By
from Backend.ScraperPrototype import ScraperPrototype
import undetected_chromedriver as uc
import img2pdf
from openpyxl.styles import PatternFill


class EscortalligatorScraper(ScraperPrototype):
    def __init__(self):
        super().__init__()
        self.path = None
        self.driver = None

        self.cities = [
            "daytona",
            "fort lauderdale",
            "fort myers",
            "gainesville",
            "jacksonville",
            "keys",
            "miami",
            "ocala",
            "okaloosa",
            "orlando",
            "palm bay",
            "panama city",
            "pensacola",
            "bradenton",
            "space coast",
            "st. augustine",
            "tallahassee",
            "tampa",
            "treasure coast",
            "west palm beach",
            "jacksonville"
        ]
        self.city = ''
        self.state = 'florida'
        self.url = ''
        self.known_payment_methods = ['cashapp', 'venmo', 'zelle', 'crypto', 'western union', 'no deposit',
                                      'deposit', ' cc ', 'card', 'credit card', 'applepay', 'donation', 'cash', 'visa',
                                      'paypal', ' mc ', 'mastercard']

        self.known_social_media = ['instagram', ' ig ', 'insta', 'snapchat', ' sc ', 'snap', 'onlyfans', 'only fans',
                                   'twitter', 'kik', 'skype', 'facebook', ' fb ', 'whatsapp', 'telegram',
                                   ' tg ', 'tiktok', 'tik tok']

        self.date_time = None
        self.scraper_directory = None
        self.screenshot_directory = None
        self.pdf_filename = None
        self.keywords = ""
        self.flagged_keywords = None
        self.only_posts_with_payment_methods = False
        self.completed = False

        self.join_keywords = False
        self.search_mode = False

        self.number_of_keywords_in_post = 0
        self.keywords_found_in_post = []

        # lists to store data and then send to Excel file
        self.phone_number = []
        self.description = []
        self.location_and_age = []
        self.links = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.timestamps = []
        self.age = []
        self.locationSplits = []

        self.number_of_keywords_found = []
        self.keywords_found = []
        self.social_media_found = []

    '''
    ---------------------------------------
    Set Data
    ---------------------------------------
    '''
    def get_cities(self) -> list:
        return self.cities

    def set_city(self, city) -> None:
        self.city = city.replace(' ', '').replace('.', '')

    def set_join_keywords(self) -> None:
        self.join_keywords = True

    def set_only_posts_with_payment_methods(self) -> None:
        self.only_posts_with_payment_methods = True

    def set_path(self, path) -> None:
        self.path = path

    def set_search_mode(self, search_mode) -> None:
        self.search_mode = search_mode

    def set_flagged_keywords(self, flagged_keywords) -> None:
        self.flagged_keywords = flagged_keywords

    def set_keywords(self, keywords) -> None:
        self.keywords = keywords

    '''
    ---------------------------------------
    Managing Scraper Run Time
    ---------------------------------------
    '''

    def initialize(self) -> None:
        # set keywords value
        #self.keywords = keywords

        self.date_time = str(datetime.today())[0:19].replace(' ', '_').replace(':', '-')

        # Format website URL based on state and city
        self.get_formatted_url()

        # Selenium Web Driver setup
        options = uc.ChromeOptions()
        # TODO - uncomment this to run headless
        if self.search_mode:
            options.add_argument('--headless=new')  # This allows the code to run without opening up a new Chrome window
        # options.headless = self.search_mode  # This determines if you program runs headless or not
        self.driver = uc.Chrome(subprocess=True, options=options)

        # Open Webpage with URL
        self.open_webpage()
        # Find links of posts
        links = self.get_links()

        # Create directory for search data
        self.scraper_directory = f'{self.path}/escortalligator-{self.city}-{self.date_time}'
        os.mkdir(self.scraper_directory)

        # Create directory for search screenshots
        self.screenshot_directory = f'{self.scraper_directory}/screenshots'
        self.pdf_filename = f'{self.screenshot_directory}/escortalligator-{self.city}-{self.date_time}.pdf'
        os.mkdir(self.screenshot_directory)

        # Get data from posts
        self.get_data(links)
        self.close_webpage()
        self.reset_variables()

    def stop_scraper(self) -> None:
        if self.search_mode:
            self.driver.close()
            self.driver.quit()
        else:
            self.driver.close()

    def open_webpage(self) -> None:
        self.driver.implicitly_wait(10)
        self.driver.get(self.url)
        self.driver.maximize_window()
        assert "Page not found" not in self.driver.page_source

    def close_webpage(self) -> None:
        self.driver.close()

    '''
    ---------------------------------------
    Getting the Data Running the Appending Functions and Getters
    ---------------------------------------
    '''
    def get_links(self) -> list:
        # click on terms btn
        btn = self.driver.find_element(
            By.CLASS_NAME, 'button')
        btn.click()

        time.sleep(2)
        # click on 2nd terms btn
        btn = self.driver.find_element(
            By.CLASS_NAME, 'footer')
        btn.click()

        posts = self.driver.find_elements(
            By.CSS_SELECTOR, '#list [href]')
        links = [post.get_attribute('href') for post in posts]
        return links[::3]

    def get_formatted_url(self) -> None:
        if self.city == 'fortlauderdale':
            self.city = 'ftlauderdale'
        self.url = f'https://escortalligator.com.listcrawler.eu/brief/escorts/usa/{self.state}/{self.city}/1'

    def get_data(self, links) -> None:
        links = set(links)
        counter = 0

        for link in links:
            self.driver.get(link)
            assert "Page not found" not in self.driver.page_source


            try:
                timestamp = self.driver.find_element(
                    By.CLASS_NAME, 'postCreatedOn').text
            except NoSuchElementException:
                timestamp = 'N/A'

            try:
                description = self.driver.find_element(
                    By.CLASS_NAME, 'viewpostbody').text
            except NoSuchElementException:
                description = 'N/A'

            try:
                phone_number = self.driver.find_element(
                    By.CLASS_NAME, 'userInfoContainer').text
            except NoSuchElementException:
                phone_number = 'N/A'

            try:
                location_and_age = self.driver.find_element(
                    By.CLASS_NAME, 'viewpostlocationIconBabylon').text
                age, locationSplits = self.parse_location_and_age(location_and_age)
            except NoSuchElementException:
                location_and_age = 'N/A'
                age = 'N/A'
                locationSplits = 'N/A'

            # reassign variables for each post
            self.number_of_keywords_in_post = 0
            self.keywords_found_in_post = []

            if self.join_keywords and self.only_posts_with_payment_methods:
                if self.check_keywords(phone_number) or self.check_keywords(location_and_age) or \
                        self.check_keywords(description):
                    counter = self.join_with_payment_methods(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)

            elif self.join_keywords or self.only_posts_with_payment_methods:
                if self.join_keywords:
                    if self.check_keywords(phone_number) or self.check_keywords(location_and_age)  or \
                            self.check_keywords(description):
                        self.check_keywords_found(description, location_and_age, locationSplits, age, phone_number, link)
                        counter = self.join_inclusive(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)

                elif self.only_posts_with_payment_methods:
                    if len(self.keywords) > 0:
                        if self.check_keywords(phone_number) or self.check_keywords(location_and_age) or \
                                self.check_keywords(description):
                            self.check_keywords_found(description, location_and_age, locationSplits, age, phone_number, link)

                    counter = self.payment_methods_only(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
            else:
                if len(self.keywords) > 0:
                    if self.check_keywords(phone_number) or self.check_keywords(location_and_age) or \
                            self.check_keywords(description):
                        self.check_keywords_found(description, location_and_age, locationSplits, age, phone_number, link)
                        self.append_data(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
                        screenshot_name = str(counter) + ".png"
                        self.capture_screenshot(screenshot_name)
                        counter += 1
                else:
                    self.append_data(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
                    screenshot_name = str(counter) + ".png"
                    self.capture_screenshot(screenshot_name)
                    counter += 1

            self.RAW_format_data_to_excel()
            self.CLEAN_format_data_to_excel()

    '''
    --------------------------
    Appending Data
    --------------------------
    '''
    def append_data(self, counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp) -> None:
        self.post_identifier.append(counter)
        self.phone_number.append(phone_number)
        self.links.append(link)
        self.location_and_age.append(location_and_age)
        self.locationSplits.append(locationSplits)
        self.age.append(age)
        self.description.append(description)
        self.check_and_append_payment_methods(description)
        self.keywords_found.append(', '.join(self.keywords_found_in_post) or 'N/A')
        self.number_of_keywords_found.append(self.number_of_keywords_in_post or 'N/A')
        self.check_for_social_media(description)
        self.timestamps.append(timestamp)

    def join_inclusive(self, counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp):
        if len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def payment_methods_only(self, counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp):
        if self.check_for_payment_methods(description):
            self.append_data(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    def join_with_payment_methods(self, counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp):
        if self.check_for_payment_methods(description) and len(self.keywords) == len(set(self.keywords_found_in_post)):
            self.append_data(counter, description, link, location_and_age, locationSplits, age, phone_number, timestamp)
            screenshot_name = str(counter) + ".png"
            self.capture_screenshot(screenshot_name)

            return counter + 1
        return counter

    '''
    --------------------------
    Checking and Running Append
    --------------------------
    '''
    def check_keywords_found(self, description, location_and_age, locationSplits, age, phone_number, link):  # add link as well?
        self.check_and_append_keywords(description)
        self.check_and_append_keywords(location_and_age)
        self.check_and_append_keywords(locationSplits)
        self.check_and_append_keywords(age)
        self.check_and_append_keywords(phone_number)
        self.check_and_append_keywords(link)

    def check_for_payment_methods(self, description) -> bool:
        for payment in self.known_payment_methods:
            if payment in description.lower():
                return True
        return False

    def check_and_append_payment_methods(self, description):
        payments = ''
        for payment in self.known_payment_methods:
            if payment in description.lower():
                payments += payment + '\n'

        if payments != '':
            self.payment_methods_found.append(payments)
        else:
            self.payment_methods_found.append('N/A')

    def check_for_social_media(self, description) -> None:
        social_media = ''
        for social in self.known_social_media:
            if social in description.lower():
                social_media += social + '\n'

        if social_media != '':
            self.social_media_found.append(social_media)
        else:
            self.social_media_found.append('N/A')

    def check_keywords(self, data) -> bool:
        for key in self.keywords:
            if key in data:
                return True
        return False

    def check_and_append_keywords(self, data) -> None:
        for key in self.keywords:
            if key in data.lower():
                self.keywords_found_in_post.append(key)
                self.number_of_keywords_in_post += 1

    '''
    ---------------------------------
    Formatting Data and Result Creation
    ---------------------------------
    '''
    def RAW_format_data_to_excel(self) -> None:
        titled_columns = {
            'Post-identifier': self.post_identifier,
            'Link': self.links,
            # -------
            'Inputted City / Region': self.city,
            'Specified Location': self.locationSplits,
            # ------
            'Timestamp': self.timestamps,
            # -------
            'Phone-Number': self.phone_number,
            # ---------
            'Age': self.age,
            # ---------
            'Description': self.description,
            # -------
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            # ---------
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        }
        # count = 2
        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/RAW-escortalligator-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(
                    i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    #call it
    def CLEAN_format_data_to_excel(self) -> None:

        # append certain info together
        contact_info = [
            f"{phone_number}"
            for phone_number in zip(
                self.phone_number,
            )
        ]
        location = [
            f"{city}||| {local}"
            for city, local in zip(
                self.city, self.locationSplits
            )
        ]

        # define columns
        titled_columns = pd.DataFrame({
            # ---- abs identifiers
            'Post-identifier': self.post_identifier,
            'Link': self.links,  # could also be a keyword source too
            # ------- time and place of posting
            'Inputted City / Region': self.city,
            'Specified Location': self.locationSplits,  # could also be a keyword source too
            'Timeline': self.timestamps,
            # ------ methods of tracking
            'Contacts': contact_info,  # could also be a keyword source too
            # ----- keyword sources
            'Personal Info': self.age,
            'Overall Description': self.description,
            # ------- other forms of transactions and communication
            'Payment-methods': self.payment_methods_found,
            'Social-media-found': self.social_media_found,
            # ------- keyword stats
            'Keywords-found': self.keywords_found,
            'Number-of-keywords-found': self.number_of_keywords_found
        })

        data = pd.DataFrame(titled_columns)
        with pd.ExcelWriter(
                f'{self.scraper_directory}/CLEAN-escortalligator-{self.city}-{self.date_time}.xlsx',
                engine='openpyxl') as writer:
            data.to_excel(writer, index=False)
            worksheet = writer.sheets['Sheet1']
            for i in range(2, worksheet.max_row):
                keywords = worksheet["K" + str(i)].value  # set the keywords var to each keyword in the cell
                for flagged_keyword in self.flagged_keywords:
                    if flagged_keyword in keywords:
                        print("flagging keyword: ", flagged_keyword)
                        worksheet["K" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')
                        worksheet["A" + str(i)].fill = PatternFill(
                            fill_type='solid',
                            start_color='ff0000',
                            end_color='ff0000')

            for col in worksheet.columns:  # dynamically adjust column sizes based on content of cell
                max_length = 0
                col = [cell for cell in col]
                for cell in col:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[
                    col[0].column_letter].width = adjusted_width

    def parse_location_and_age(self, location_and_age_str):
        # Example input: "Age: 24Location:  Incall and outcall Daytona"
        age = 'N/A'
        location = 'N/A'

        # Splitting the string based on 'Location:' to separate age and location
        parts = location_and_age_str.split('Location:')

        # Extracting age part and removing 'Age:' prefix
        if len(parts) > 0:
            age_part = parts[0].strip()
            age = age_part.replace('Age:', '').strip()

        # Extracting location if present
        if len(parts) > 1:
            location = parts[1].strip()

        return age, location

    def capture_screenshot(self, screenshot_name) -> None:
        self.driver.save_screenshot(f'{self.screenshot_directory}/{screenshot_name}')
        self.create_pdf()

    def create_pdf(self) -> None:
        screenshot_files = [
            os.path.join(self.screenshot_directory, filename) for filename in os.listdir(self.screenshot_directory) if
            filename.endswith('.png')]
        with open(self.pdf_filename, "wb") as f:
            f.write(img2pdf.convert(screenshot_files))

    def reset_variables(self) -> None:
        self.phone_number = []
        self.description = []
        self.locationSplits = []
        self.age = []
        self.location_and_age = []
        self.links = []
        self.post_identifier = []
        self.payment_methods_found = []
        self.number_of_keywords_found = []
        self.keywords_found = []
        self.only_posts_with_payment_methods = False
        self.join_keywords = False
        self.social_media_found = []
